from noise_reduction import image_preprocessing
from PIL import Image
import fitz 
from openpyxl import load_workbook, Workbook
import io
import os 

_FONT_QN = ['PalatinoLinotype-Roman', 'PalatinoLinotype-Bold']

def save_image(
    img: Image,
    save_path,
    have_image_preprocessing: bool = 0,
) -> None:
    ''' luu anh voi tien xu ly anh 

    args: 
        img: du lieu anh
        save_path: duong dan luu anh
        have_image_preprocessing: co tien xu ly hinh anh truoc khi luu khong ? 
    '''
    if have_image_preprocessing:
        img = image_preprocessing(img)

    img.save(save_path)

def extract_image(
    pdf_file_name: str, page_index: int
) -> tuple:
    ''' trich xuat du lieu hinh anh 

    args: 
        pdf_file_name: ten file pdf can trich xuat 
        page_index: vi tri trang can trich xuat
    returns:
        (img, ext) = (du lieu hinh anh, duoi file anh) 
    '''

    pdf_file = fitz.open(pdf_file_name)
    
    images = pdf_file.load_page(page_index).get_images()

    xref = images[0][0]
    base_image = pdf_file.extract_image(xref) # extract images information (width, height, .....)

    ext = base_image['ext'] # image file extension (ex: 'ext': 'png')
        
    img = Image.open(io.BytesIO(base_image['image']))

    return img, ext
 
def extract_text(
    pdf_file_name: str, 
    page_index: int
)-> tuple:
    ''' quote but text is QN and HN in pdf file

    args: 
        pdf_file_name: ten file pdf de lay text 
        page_index: vi tri trang can trich xuat 

    returns:
        List of sentences in Vietnamese and Han Nom 
        searched on the page
    '''

    pdf_file = fitz.open(pdf_file_name)
    load_page = pdf_file.load_page(page_index)

    dictionary_elements = load_page.get_text('dict') 
    
    QN = HN = ''
    list_HN, list_QN = [], [] 

    for block in dictionary_elements['blocks']:
        for line in block['lines']:
            for span in line['spans']:
                
                # Based on the font specification 
                # to determine whether the line is Vietnamese or Nom
                if span['font'] in _FONT_QN: 

                    HN = ' '.join(HN.split())
                    
                    # Nom language omits spaces  
                    if HN: 
                        HN = HN.replace(' ', '')
                        HN = HN.replace('。。', '。')
                        list_HN.append(HN) 
                    HN = ''

                    QN += ' ' + span['text']
                
                else:
                    
                    QN = ' '.join(QN.split())
                    if QN: list_QN.append(QN.replace('  ', ' ')) 
                    QN = ''

                    HN += span['text']

    HN, QN = ' '.join(HN.split()), ' '.join(QN.split())

    list_QN.append(HN.replace('。。', '。') ) if HN else ''
    list_QN.append(QN) if QN else ''
    
    return list_HN, list_QN

def print_xlsx(
    file_name: str, 
    start_row: int, 
    sentences_HN: list[str], 
    sentences_QN: list[str], 
    page_index: int
) -> int:
    ''' print sentences from pdf file to xlsx file

    args: 
        file_name: fil excel name
        start_row: start row to print content 
        sentences_(HN, QN): a list text sentences in file pdf
        page_index: determine which index page is being processed
    
    returns: 
        end row position in xlsx file
    '''
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    worksheet = workbook.active 
    
    worksheet.cell(1, 1, 'Page Index')
    worksheet.cell(1, 2, 'SinoNom Char')
    worksheet.cell(1, 3, 'Chữ Quốc Ngữ')

    page_index += 1
    row = start_row
    for QN in sentences_QN:
        worksheet.cell(row, 1, page_index)               
        worksheet.cell(row, 3, QN)

        row += 1
    
    row = start_row
    for HN in sentences_HN:
        worksheet.cell(row, 1, page_index)               
        worksheet.cell(row, 2, HN)

        row += 1

    workbook.save(file_name)
    
    return row

def process_page(
    pdf_file_name: str, 
    folder_save_image: str,
    page_index: int,
    row_xlsx: int = 1,
    have_image_preprocessing=True
) -> None:
    ''' Process the page to retrieve content and images

    args:
        pdf_file_name: ten file pdf can xu ly 
        folder_save_image: folder luu anh sau khi da trich xuat
        page_index: determine which index page is being processed
        row_xlsx: row in file xlsx
    
    returns:
        dong cuoi cung sau khi in ra file excel
    '''

    pdf_file = fitz.open(pdf_file_name)
    images = pdf_file.load_page(page_index).get_images()

    if images:
        img, ext = extract_image(pdf_file_name, page_index)
        
        # path to the storage location in the specified folder
        full_file_name = pdf_file_name[:pdf_file_name.find('.')] + '.' + str(page_index + 1) +'.' + ext 
        image_save_path = folder_save_image + full_file_name
        
        save_image(img, image_save_path, have_image_preprocessing)
    
    else:
        sentences_HN, sentences_QN = extract_text(pdf_file_name, page_index)
        
        file_output = 'text_' + pdf_file_name[:pdf_file_name.find('.')] + '.xlsx'
        
        row_xlsx = print_xlsx(file_output, row_xlsx, sentences_HN, sentences_QN, page_index)
    
    return row_xlsx

def extract_pdf(
    pdf_file_name: str, 
    folder_save_image: str
) -> None:
    ''' extract images and text from file pdf

    args:
        pdf_file_name: ten file pdf can trich xuat 
        folder_save_image: folder chua anh sau khi trich xuat duoc
    '''
    pdf_file = fitz.open(pdf_file_name)
 
    row_xlsx = 2
    for page_index in range(len(pdf_file)):
        row_xlsx = process_page(pdf_file_name, folder_save_image, page_index, row_xlsx)
    
            
if __name__ == '__main__':
    PDF_FILE_NAME = 'Bondaytrecon.pdf'
    FOLDER_SAVE_IMAGE = 'dang1/'
    
    try:
        os.mkdir(FOLDER_SAVE_IMAGE)
    except FileExistsError:
        pass

    extract_pdf(PDF_FILE_NAME, FOLDER_SAVE_IMAGE)
